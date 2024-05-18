import customtkinter
from typing import Text
import openpyxl
import time
from tqdm import tqdm
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"
from utils import load_pickle, latest_change, dump_pickle, open_file
from tkinter.filedialog import askopenfile
from toan_bo_data import ToanBoData
from nhan_vien import NhanVien
from constant import DATANAME, DATAPATH, DI_TRE, NGHI_PHEP, OUTPUT_PRINT_EXCEL, TANG_CA, TEMPLATE_EXCEL, VE_SOM
from ghi_chep_hang_ngay import GhiChepHangNgay, HanhChanh, TangCa
from cong_viec import CongViec
import os

# region hyper-parameter
STT = 0
START_ROW = 13
START_DATA = 8
# endregion hyper-parameter

class App(customtkinter.CTk):
    def __init__(
            self,
            data_path : Text = DATANAME
        ):
        super().__init__()

        # region 1: Lấy dữ liệu
        self.data : ToanBoData = ToanBoData()

        lst_nam = self.data.get_danh_sach_nam()
        # endregion


        # region 2: Configure window
        self.title("Danh Product")
        self.geometry(f"{1000}x{500}") # Chỉnh kích thước màn hình
        # endregion


        # region 3: Tạo một sidebar để chọn tháng và năm
        self.sidebar_frame = customtkinter.CTkFrame(self, width=100, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")

        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame,
            text="Tháng - năm",
            font=customtkinter.CTkFont(
                size=15,
                weight="bold"
            ),
            text_color= 'red'
        )
        self.logo_label.grid(row=0, column=0, padx=0, pady=(0, 0))

            # region Chọn năm
        self.nam_label = customtkinter.CTkLabel(self.sidebar_frame, text="Năm:", anchor="w")
        self.nam_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.nam_optionemenu = customtkinter.CTkOptionMenu(
            self.sidebar_frame, 
            values= sorted(lst_nam),
            command= self.change_month_option,
            width=50
        )
        self.nam_optionemenu.grid(row=6, column=0, padx=0, pady=(10, 20))
            # endregion

            # region Chọn tháng
        self.thang_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Chọn tháng:", anchor="w")
        self.thang_label.grid(row=7, column=0, padx=0, pady=(0, 0))

        lst_thang = self.data.get_danh_sach_thang(nam = self.nam_optionemenu.get())
        # lst_thang = sorted(list(map(lambda x: x.zfill(2), lst_thang)))
        self.thang_optionemenu = customtkinter.CTkOptionMenu(
            self.sidebar_frame, 
            values = sorted(lst_thang, key = lambda x: int(x)),
            command= self.sidebar_button_event,
            width=50
        )
        self.thang_optionemenu.grid(row=8, column=0, padx=0, pady=(0,0))
            # endregion

            # region Truy xuất dữ liệu
        # self.bt_truy_xuat = customtkinter.CTkButton(
        #     self.sidebar_frame, 
        #     command=self.sidebar_button_event,
        #     text= "Truy xuất"
        #     )
        # self.bt_truy_xuat.grid(row=9, column=0, padx=20, pady=10)
            # endregion

        # endregion
        
        
        # region 4: Create slider and progressbar frame
        self.hor_frame = customtkinter.CTkFrame(
            self,
            width=250,
            # corner_radius= 0
        )
        
        self.hor_frame.grid(
            row=0,
            column=1,
            columnspan=2, 
            padx=(0, 0), 
            # pady=(10, 0), 
            sticky="nsew",
            rowspan=1
        )
        # self.hor_frame.grid_rowconfigure(4, weight=1)

            # region 4.1: Logo Kết quả công việc
        self.logo_label = customtkinter.CTkLabel(
            self.hor_frame, 
            text=f"Kết quả làm việc {self.thang_optionemenu.get().zfill(2)}/{self.nam_optionemenu.get()}", 
            font=customtkinter.CTkFont(
                size=15,
                weight="bold",
            ),
            text_color='red'
        )
        self.logo_label.grid(row=3, column=2, padx=0, pady=(0, 0))
            # endregion

            # region 4.2: Tải danh sách nhân viên
                
                # region 4.2.1: Nhãn `Mã nhân viên``
        self.nhan_vien_label = customtkinter.CTkLabel(self.hor_frame, text="Mã nhân viên:", anchor="w")
        self.nhan_vien_label.grid(row=4, column=1, padx=0, pady=(0, 0))
                # endregion
        
                # region 4.2.2: Option Menu
        self.nhan_vien_optionemenu = customtkinter.CTkOptionMenu(
            self.hor_frame, 
            width= 50,
            values= self.data.get_danh_sach_nhan_vien(
                nam= self.nam_optionemenu.get(),
                thang= self.thang_optionemenu.get()
            )
        )
        self.nhan_vien_optionemenu.grid(row=4, column=2, padx=0, pady=(0, 0))
                # endregion

                # region 4.2.3: Xem danh sách nhân viên
        self.bt_xem_chi_tiet_cong_viec = customtkinter.CTkButton(
            self.hor_frame, 
            width=100,
            command=self.button_xem_nhan_vien,
            text= "Xem nhân viên"
        )
        self.bt_xem_chi_tiet_cong_viec.grid(row=4, column=3, padx=0, pady=0)
                # endregion
        # endregion
            
            # region Label các công việc
        self.cac_cong_viec_label = customtkinter.CTkLabel(self.hor_frame, text=f"Các công việc: ", anchor="w")
        self.cac_cong_viec_label.grid(row=8, column=1, padx=0, pady=(0, 0))
            # endregion

            # region options công việc
        self.nhan_vien = None
        self.ho_ten_nv = ""
        self.ma_so_nv = ""
        self.cac_cong_viec_trong_thang = ""
        self.so_cong_viec_trong_thang = []
        self.so_gio_cong = 0
        self.so_phut_tre = 0
        self.so_gio_ve_som = 0
        self.so_gio_tang_ca = 0
        self.so_gio_nghi_phep = 0

        # region Label Họ tên
        self.ho_ten_label = customtkinter.CTkLabel(self.hor_frame, text=f"Họ tên: {self.ho_ten_nv}", anchor="w")
        self.ho_ten_label.grid(row=5, column=1, padx=0, pady=(0, 0))
        # endregion

        # region Số ngày công
        self.so_gio_cong_label = customtkinter.CTkLabel(self.hor_frame, text=f"Số ngày công: {self.so_gio_cong} (giờ)", anchor="w")
        self.so_gio_cong_label.grid(row=6, column=1, padx=0, pady=(0, 0))
        # endregion

        # region số công việc
        self.tang_ca_label = customtkinter.CTkLabel(self.hor_frame, text=f"Tăng ca: {self.so_gio_tang_ca} (giờ)", anchor="w")
        self.tang_ca_label.grid(row=5, column=3, padx=0, pady=(0, 0))
        # endregion

        # region Trễ
        self.tre_label = customtkinter.CTkLabel(self.hor_frame, text=f"Trễ: {self.so_phut_tre} (phút)", anchor="w")
        self.tre_label.grid(row=5, column=2, padx=0, pady=(0, 0))
        # endregion

        # region Sớm
        self.som_label = customtkinter.CTkLabel(self.hor_frame, text=f"Sớm: {self.so_gio_ve_som} (giờ)", anchor="w")
        self.som_label.grid(row=6, column=2, padx=0, pady=(0, 0))
        # endregion

        # region Phép
        self.phep_label = customtkinter.CTkLabel(self.hor_frame, text=f"Phép: {self.so_gio_nghi_phep} (giờ)", anchor="w")
        self.phep_label.grid(row=6, column=3, padx=0, pady=(0, 0))
        # endregion
        lst_cong_viec = self.cac_cong_viec_trong_thang
        self.cac_cong_viec_optionemenu = customtkinter.CTkOptionMenu(
            self.hor_frame, 
            values= lst_cong_viec,
            width=50
        )
        self.cac_cong_viec_optionemenu.grid(row=8, column=2, padx=0, pady=(0, 0))
            # endregion

            # region nút "Xem chi tiết công việc"
        self.bt_xem_chi_tiet_cong_viec = customtkinter.CTkButton(
            self.hor_frame, 
            command=self.button_xem_chi_tiet,
            text= "Xem chi tiết công việc"
        )
        self.bt_xem_chi_tiet_cong_viec.grid(row=8, column=3, padx=0, pady=0)

            # endregion
        # endregion

        # region 3: Tạo một sidebar để Thêm và cập nhật dữ liệu
        self.sidebar_frame_upload = customtkinter.CTkFrame(self, width=100, corner_radius=0)
        self.sidebar_frame_upload.grid(row=0, column=3, sticky="nsew", padx = (0, 0), pady = (0, 0), rowspan=4)

            # region Logo
        self.logo_label_2 = customtkinter.CTkLabel(
            self.sidebar_frame_upload, 
            text="Upload File",
            font=customtkinter.CTkFont(
                size=15, 
                weight="bold"
            ), 
            text_color= 'red'
            )
        self.logo_label_2.grid(row=0, column=0, padx=0, pady=(0, 0))
            # endregion

            # region Cập nhật toàn bộ
        self.update_whole = customtkinter.CTkButton(
            self.sidebar_frame_upload, 
            command=self.update_whole,
            text= "Cập nhật",
            width= 50
        )
        self.update_whole.grid(row=2, column=0, padx=0, pady=30)
            # endregion

        #     # region Cập nhật 1 phần
        # self.update_partial = customtkinter.CTkButton(
        #     self.sidebar_frame_upload, 
        #     command=self.update_partial,
        #     text= "Cập nhật một phần"
        # )
        # self.update_partial.grid(row=6, column=0, padx=20, pady=10)
        #     # endregion



            # region Tải file lên
        self.upload_file = customtkinter.CTkButton(
            self.sidebar_frame_upload, 
            command=self.upload_file,
            text= "Tải file",
            width = 50
        )
        self.upload_file.grid(row=8, column=0, padx=0, pady=30)
            # endregion

        #     # region Chọn năm
        # self.nam_label = customtkinter.CTkLabel(self.sidebar_frame, text="Năm:", anchor="w")
        # self.nam_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        # self.nam_optionemenu = customtkinter.CTkOptionMenu(
        #     self.sidebar_frame, 
        #     values= sorted(lst_nam)
        # )
        # self.nam_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 20))
        #     # endregion

        #     # region Chọn tháng
        # self.thang_label = customtkinter.CTkLabel(
        #     self.sidebar_frame, text="Chọn tháng:", anchor="w")
        # self.thang_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        # self.thang_optionemenu = customtkinter.CTkOptionMenu(
        #     self.sidebar_frame, 
        #     values = sorted(self.data.get_danh_sach_thang(nam = self.nam_optionemenu.get()))
        # )
        # self.thang_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 10))
        #     # endregion

        #     # region Truy xuất dữ liệu
        # self.bt_truy_xuat = customtkinter.CTkButton(
        #     self.sidebar_frame, 
        #     command=self.sidebar_button_event,
        #     text= "Truy xuất"
        #     )
        # self.bt_truy_xuat.grid(row=9, column=0, padx=20, pady=10)
        #     # endregion

        # endregion
        
        
        # region Khu vực để hiển thị nội dung chi tiết cho 1 người
        self.textbox = customtkinter.CTkTextbox(self, width=250)
        self.textbox.grid(row=1, column=1, padx=(0, 0), pady=(0, 0), sticky="nsew", columnspan = 2)

        # endregion



        # region 4: Tạo vung in file ra excel
        self.sidebar_frame_print = customtkinter.CTkFrame(self, width=100, corner_radius=0)
        self.sidebar_frame_print.grid(row=0, column=4, sticky="nsew", padx = (0, 0), pady = (0, 0), rowspan=2)

            # region Logo
        self.logo_label_print = customtkinter.CTkLabel(
            self.sidebar_frame_print, 
            text="In ra Excel", 
            font=customtkinter.CTkFont(
                size=15, 
                weight="bold"
                ), 
            text_color='red')
        self.logo_label_print.grid(row=0, column=0, padx=0, pady=(0, 0))
            # endregion
        lst_checkbox = []
        lst_nhan_vien = self.data.get_danh_sach_nhan_vien(
                nam= self.nam_optionemenu.get(),
                thang= self.thang_optionemenu.get()
            )
        self.dict_nhan_vien_checkbox = {}
        for idx, x in enumerate(lst_nhan_vien, start= 1):
            self.dict_nhan_vien_checkbox[x] = customtkinter.CTkCheckBox(master=self.sidebar_frame_print, text = x  )
            if idx % 2 == 0: #chan
                self.dict_nhan_vien_checkbox[x].grid(row=idx, column=0, pady=(0, 0), padx=0, sticky="n")
            else:
                self.dict_nhan_vien_checkbox[x].grid(row=idx+1, column=1, pady=(0, 0), padx=0, sticky="n")
        
        self.bt_in_ra_excel = customtkinter.CTkButton(
            self.sidebar_frame_print, 
            command=self.print_to_excel,
            text= "In ra excel",
            width = 70
            )
        self.bt_in_ra_excel.grid(row=len(lst_nhan_vien)+1, column=0, padx=20, pady=10)

        self.logo_output_name = customtkinter.CTkLabel(
            self.sidebar_frame_print,
            text="Đầu ra",
            font=customtkinter.CTkFont(
                size=13,
                weight="bold"
            )
        )
        self.logo_output_name.grid(row=len(lst_nhan_vien)+2, column=0, padx=0, pady=(0, 0))


        self.logo_progressbar_print_1 = customtkinter.CTkLabel(
            self.sidebar_frame_print,
            text="Tiến độ ghi vào file excel:",
            font=customtkinter.CTkFont(
                size=13,
                weight="bold"
            )
        )
        self.logo_progressbar_print_1.grid(row=len(lst_nhan_vien)+3, column=0, padx=0, pady=(0, 0))


        self.logo_progressbar_print_2 = customtkinter.CTkLabel(
            self.sidebar_frame_print,
            text="Tiến độ ghép các ô:",
            font=customtkinter.CTkFont(
                size=13,
                weight="bold"
            )
        )
        self.logo_progressbar_print_2.grid(row=len(lst_nhan_vien)+5, column=0, padx=0, pady=(0, 0))


        self.logo_progressbar_print_3 = customtkinter.CTkLabel(
            self.sidebar_frame_print,
            text="Tiến độ chuẩn hóa ô:",
            font=customtkinter.CTkFont(
                size=13,
                weight="bold"
            )
        )
        self.logo_progressbar_print_3.grid(row=len(lst_nhan_vien)+7, column=0, padx=0, pady=(0, 0))



        self.progressbar_print = customtkinter.CTkProgressBar(self.sidebar_frame_print, orientation="horizontal")
        self.progressbar_print.grid(row=len(lst_nhan_vien)+4, column=0, padx=0, pady=(0, 0))

        self.progressbar_print_2 = customtkinter.CTkProgressBar(self.sidebar_frame_print, orientation="horizontal")
        self.progressbar_print_2.grid(row=len(lst_nhan_vien)+6, column=0, padx=0, pady=(0, 0))

        self.progressbar_print_3 = customtkinter.CTkProgressBar(self.sidebar_frame_print, orientation="horizontal")
        self.progressbar_print_3.grid(row=len(lst_nhan_vien)+8, column=0, padx=0, pady=(0, 0))

        self.progressbar_print.set(0)
        self.progressbar_print_2.set(0)
        self.progressbar_print_3.set(0)

        # # region Khung tìm kiếm
        # self.entry = customtkinter.CTkEntry(self, placeholder_text="CTkEntry")
        # self.entry.grid(row=8, column=1, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")

        # self.main_button_1 = customtkinter.CTkButton(master=self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        # self.main_button_1.grid(row=9, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        # # endregion


        # # region create tabview
        # self.tabview = customtkinter.CTkTabview(self, width=250)
        # self.tabview.grid(row=0, column=3, padx=(20, 0), pady=(20, 0), sticky="nsew")
        # self.tabview.add("CTkTabview")
        # self.tabview.add("Tab 2")
        # self.tabview.add("Tab 3")
        # self.tabview.tab("CTkTabview").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        # self.tabview.tab("Tab 2").grid_columnconfigure(0, weight=1)

        # self.optionmenu_1 = customtkinter.CTkOptionMenu(self.tabview.tab("CTkTabview"), dynamic_resizing=False,
        #                                                 values=["Value 1", "Value 2", "Value Long Long Long"])
        # self.optionmenu_1.grid(row=0, column=3, padx=20, pady=(20, 10))
        # self.combobox_1 = customtkinter.CTkComboBox(self.tabview.tab("CTkTabview"),
        #                                             values=["Value 1", "Value 2", "Value Long....."])
        # self.combobox_1.grid(row=1, column=3, padx=20, pady=(10, 10))
        # self.string_input_button = customtkinter.CTkButton(self.tabview.tab("CTkTabview"), text="Open CTkInputDialog",
        #                                                    command=self.open_input_dialog_event)
        # self.string_input_button.grid(row=2, column=3, padx=20, pady=(10, 10))
        # self.label_tab_2 = customtkinter.CTkLabel(self.tabview.tab("Tab 2"), text="CTkLabel on Tab 2")
        # self.label_tab_2.grid(row=0, column=3, padx=20, pady=20)

        # # endregion

    def print_to_excel(self):
        # Lay du lieu
        self.progressbar_print.set(0)
        self.progressbar_print_2.set(0)
        self.progressbar_print_3.set(0)

        year = self.nam_optionemenu.get()
        month = self.thang_optionemenu.get()
        dict_nhan_vien = self.data.dict_ghi_chep_cac_nam[year].dict_ghi_chep_12_thang[month].dict_nhan_vien

        # Lay cac checkbox duoc check
        lst_nhan_vien_need_print = []
        for k, v in self.dict_nhan_vien_checkbox.items():
            if v.get():
                lst_nhan_vien_need_print.append(k)

        # region Load tempalte
        workbook = openpyxl.load_workbook(TEMPLATE_EXCEL)
        sheet = workbook.active
        # endregion

        # region Change title 
        sheet.cell(row = 1,column = 1).value = f'KẾ HOẠCH VÀ TIẾN ĐỘ PHAY THÁNG {str(month).zfill(2)} NĂM {year}'
        cell = sheet.cell(row = 1,column = 1)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial", size=22, bold= True)
        # endregion

        # region hyper-parameter
        STT = 0
        START_ROW = 13
        START_DATA = 8
        # endregion hyper-parameter

        # region file out

        if not os.path.exists(OUTPUT_PRINT_EXCEL):
            os.makedirs(name = OUTPUT_PRINT_EXCEL, exist_ok= True)

        year_fd = os.path.join(OUTPUT_PRINT_EXCEL, f'Y-{year}')
        if not os.path.exists(year_fd):
            os.makedirs(name = year_fd, exist_ok= True)
        str_lst_nhan_vien = '_'.join(lst_nhan_vien_need_print)
        file_out = os.path.join(year_fd,f'T-{month}_{str_lst_nhan_vien}.xlsx')
        # endregion

        print('Step 2: Write data to excel')
        row_cur = START_ROW
        s2 = time.perf_counter()
        
        self.progressbar_print.start()
        for idx, ma_nhan_vien in tqdm(enumerate(lst_nhan_vien_need_print), desc = "Process to write data to excel"):
            value = idx/len(lst_nhan_vien_need_print)
            self.progressbar_print.set(value)
            tat_ca_chi_tiet = list(dict_nhan_vien[ma_nhan_vien].dict_ghi_chep_cong_viec.keys())
            for chi_tiet in tat_ca_chi_tiet:
                tat_ca_ban_ve = list(dict_nhan_vien[ma_nhan_vien].dict_ghi_chep_cong_viec[chi_tiet].keys())
                for ban_ve in tat_ca_ban_ve:
                    cong_viec = dict_nhan_vien[ma_nhan_vien].dict_ghi_chep_cong_viec[chi_tiet][ban_ve][0]

                    cong_doan = cong_viec.cong_doan
                    nha_may = cong_viec.nha_may
                    so_luong = cong_viec.so_luong
                    don_vi = cong_viec.don_vi

                    so_ngay_lam = len(cong_viec.ghi_ghep_hang_ngay)
                    lst_ngay_cong = [cong_viec.ghi_ghep_hang_ngay[ngay_lam].ngay for ngay_lam in range(so_ngay_lam)]
                    STT+=1

                    sheet.cell(row = row_cur,column = 1).value = STT
                    sheet.cell(row = row_cur,column = 2).value = chi_tiet
                    sheet.cell(row = row_cur,column = 3).value = ban_ve
                    sheet.cell(row = row_cur,column = 4).value = cong_doan
                    sheet.cell(row = row_cur,column = 5).value = nha_may
                    sheet.cell(row = row_cur,column = 6).value = 'KH'
                    sheet.cell(row = row_cur,column = 7).value = so_luong
                    sheet.cell(row = row_cur,column = 8).value = don_vi
                    sheet.cell(row = row_cur,column = 40).value = ma_nhan_vien

                    sheet.cell(row = row_cur + 1 ,column = 6).value = 'TT'
                    sheet.cell(row = row_cur + 1,column = 7).value = so_luong

                    # Get workday
                    for ngay_cong in lst_ngay_cong:
                        sheet.cell(row = row_cur,column = START_DATA + int(str(ngay_cong))).fill = PatternFill(start_color= "008080", end_color="008080", fill_type="solid")
                        sheet.cell(row = row_cur+1,column = START_DATA + int(str(ngay_cong))).fill = PatternFill(start_color= "00FF00", end_color="00FF00", fill_type="solid")

                    
                    row_cur+=2
        self.progressbar_print.stop()
        
        e2 = time.perf_counter()
        print(f'Time write data to excel: {e2 - s2:.2f}s')
        # Merge cell
        print(f"Step 3: Merge cell with column 1,2,3,4,5,8 and 40")
        lst_merge_col = [1,2,3,4,5,8, 40]
        
        self.progressbar_print_2.start()
        for row in tqdm(range(START_ROW, row_cur, 2), desc = 'Process to Merge cell'):
            self.progressbar_print_2.set(row/row_cur)
            for col in lst_merge_col:
                sheet.merge_cells(start_row=row, end_row=row+1, start_column=col, end_column=col)

        self.progressbar_print_2.stop()
        # Format the cells
        print(f"Step 4: Format the cells")
        
        self.progressbar_print_3.start()
        idx = 0
        for row in tqdm(sheet.iter_rows(min_row= START_ROW, min_col=1, max_col = 40), desc= "Process to Format the cells"):
            idx+=1
            self.progressbar_print_3.set(idx/row_cur)
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=11)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.progressbar_print_3.stop()
        # region save the output file
        print("Step 5: Save output file")
        print(f"****OUTPUT FILE: {file_out}****")
        s3 = time.perf_counter()
        workbook.save(file_out)
        e3 = time.perf_counter()
        print(f'Time save output file: {e3 - s3:.2f}s')
        print("================================")
        # endregion

        self.logo_output_name.configure(text = f'{file_out}')

    # endregion
        


    def sidebar_button_event(self, thang):
        # region Logo Kết quả công việc
        self.logo_label.configure(text = f"Kết quả làm việc {thang.zfill(2)}/{self.nam_optionemenu.get()}")
        # endregion

        # region Tải danh sách nhân viên
        self.nhan_vien_optionemenu.configure(
            values= self.data.get_danh_sach_nhan_vien(
                nam= self.nam_optionemenu.get(),
                thang= self.thang_optionemenu.get())
        )
        # endregion


        # region Lấy danh sách công việc của nhân viên đầu tiên để hiển thị
        self.nhan_vien : NhanVien = self.data.get_thong_tin_nhan_vien(
            nam = self.nam_optionemenu.get(),
            thang= self.thang_optionemenu.get(),
            ma_nhan_vien= self.nhan_vien_optionemenu.get()
        )
        self.ho_ten_nv = self.nhan_vien.get_ho_ten()
        self.ma_so_nv = self.nhan_vien.get_msnv()
        self.cac_cong_viec_trong_thang = self.nhan_vien.get_cac_cong_viec_trong_thang()
        self.so_cong_viec_trong_thang = self.nhan_vien.get_so_cong_viec_trong_thang()
        self.so_gio_cong = 30
        self.so_phut_tre = self.nhan_vien.get_so_phut_di_tre()
        self.so_gio_ve_som, *_ = self.nhan_vien.get_so_gio_ve_som()
        self.so_gio_tang_ca, *_ = self.nhan_vien.get_so_gio_tang_ca()
        self.so_gio_nghi_phep, *_ = self.nhan_vien.get_so_gio_nghi_phep()

        self.cac_cong_viec_optionemenu.configure(
            values = self.cac_cong_viec_trong_thang
        )
        # endregion

    def button_xem_chi_tiet(self):
        text = ''
        cv = self.cac_cong_viec_optionemenu.get()
        ban_ve = list(self.nhan_vien.dict_ghi_chep_cong_viec[cv].keys())
        text+=f"***** {cv} ****\n"
        lst_ngay_tre = []
        lst_so_phut_tre = []
        if cv == DI_TRE:
            for bv in ban_ve:
                cong_viec : CongViec = self.nhan_vien.dict_ghi_chep_cong_viec[cv][bv][0]
                # Ngày trễ
                ngay_tre = str(cong_viec.ghi_ghep_hang_ngay[0].ngay).zfill(2)
                lst_ngay_tre.append(ngay_tre)

                # Số phút trễ
                so_phut_tre = cong_viec.ghi_ghep_hang_ngay[0].hanh_chanh.TSN_nghi
                lst_so_phut_tre.append(so_phut_tre)

            for idx, (ngay, phut) in enumerate(zip(lst_ngay_tre, lst_so_phut_tre)):
                text+=f'''
{idx+1}. Ngày {ngay} trễ {phut} phút.
'''     
        elif cv == TANG_CA:
            (
                so_gio_tang_ca,
                lst_ngay_tang_ca,
                lst_gio_tang_ca,
                lst_cong_viec,
                lst_ban_ve
            )= self.nhan_vien.get_so_gio_tang_ca()
            for idx, (ngay, gio, _cong_viec, _ban_ve) in enumerate(zip(lst_ngay_tang_ca, lst_gio_tang_ca, lst_cong_viec, lst_ban_ve)):
                text+=f'''
{idx+1}. Ngày {ngay} tăng ca {gio} giờ cho công việc {_cong_viec} - bản vẽ {_ban_ve}.
'''         
            text+=f'\n==> Tổng cộng tăng ca: {so_gio_tang_ca} (giờ)'
        elif cv == VE_SOM:
            (
                so_gio_ve_som,
                lst_ngay_ve_som,
                lst_gio_ve_som,
            )= self.nhan_vien.get_so_gio_ve_som()
            for idx, (ngay, gio) in enumerate(zip(lst_ngay_ve_som, lst_gio_ve_som)):
                text+=f'''
{idx+1}. Ngày {ngay} về sớm {gio} giờ.
'''         
            text+=f'\n==> Tổng cộng về sớm: {so_gio_ve_som} (phút)'
        elif cv == NGHI_PHEP:
            (
                so_gio_nghi_phep,
                lst_ngay_nghi_phep,
                lst_gio_nghi_phep,
            )= self.nhan_vien.get_so_gio_nghi_phep()
            for idx, (ngay, gio) in enumerate(zip(lst_ngay_nghi_phep, lst_gio_nghi_phep)):
                text+=f'''
{idx+1}. Ngày {ngay} nghỉ phép {gio} giờ.
'''         
            text+=f'\n==> Tổng cộng nghỉ phép: {so_gio_nghi_phep} (phút)'
        else:
            for bv in ban_ve:
                cong_viec = self.nhan_vien.dict_ghi_chep_cong_viec[cv][bv][0]
                cong_doan = cong_viec.cong_doan
                nha_may = cong_viec.nha_may
                # so_luong = cong_viec.so_luong
                # don_vi = cong_viec.don_vi
                kh_dt = cong_viec.KH_DT_CT
                pyc = cong_viec.PYC
                so_ngay_lam = len(cong_viec.ghi_ghep_hang_ngay)
                lst_ngay_cong = [cong_viec.ghi_ghep_hang_ngay[ngay_lam].ngay for ngay_lam in range(so_ngay_lam)]
                lst_gio_hanh_chanh = []
                lst_ti_le_hanh_chanh = []
                lst_tsn_nghi_hanh_chanh = []

                lst_gio_tang_ca = []
                lst_ti_le_tang_ca = [] 
                for ngay_lam in range(so_ngay_lam):
                    hanh_chinh : HanhChanh = cong_viec.ghi_ghep_hang_ngay[ngay_lam].hanh_chanh
                    tang_ca : TangCa = cong_viec.ghi_ghep_hang_ngay[ngay_lam].tang_ca
                    if tang_ca is not None:
                        # Tang ca
                        lst_gio_tang_ca.append(tang_ca.gio)
                        lst_ti_le_tang_ca.append(tang_ca.ti_le)
                        
                    if hanh_chinh is not None:
                        # Hanh chanh
                        lst_gio_hanh_chanh.append(hanh_chinh.gio)
                        lst_ti_le_hanh_chanh.append(hanh_chinh.ti_le)
                        lst_tsn_nghi_hanh_chanh.append(hanh_chinh.TSN_nghi)
                
                ti_le = 0
                print(f'lst_ti_le_tang_ca : {lst_ti_le_tang_ca}')
                lst_ti_le_tang_ca = list(map(lambda x: 0 if x is None else x,  lst_ti_le_tang_ca))
                if lst_ti_le_tang_ca is not None:
                    lst_ti_le_hanh_chanh.extend(lst_ti_le_tang_ca)

                print(f'lst_ti_le_hanh_chanh : {lst_ti_le_hanh_chanh}')
                lst_ti_le_hanh_chanh = list(map(lambda x: 0 if x is None else x,  lst_ti_le_hanh_chanh))
                if lst_ti_le_hanh_chanh :
                    ti_le = max(lst_ti_le_hanh_chanh)*100

                print(f'lst_gio_hanh_chanh: {lst_gio_hanh_chanh}')
                print(f'lst_gio_hanh_chanh: {lst_gio_tang_ca}')
                tong_gio_hc = sum(lst_gio_hanh_chanh)
                tong_gio_tc = sum(lst_gio_tang_ca)
                    

                text+=(f"""
1. Bản vẽ: {bv}
2. Công đoạn: {cong_doan}
3. Nhà máy : {nha_may}
4. Số giờ làm hành chính: {tong_gio_hc} (giờ)
5. Số giờ tăng ca: {tong_gio_tc} (giờ)
6. Tổng số giờ làm: {tong_gio_hc + tong_gio_tc} (giờ)
7. Tỉ lệ hoàn thành: {ti_le} (%)
8: Danh sách ngày làm: {', '.join(lst_ngay_cong)}
9. KH/ĐT: {kh_dt}
10. Phiếu yêu cầu: {pyc}
-------------------------------------------------------
    """)
        # Xóa toàn bộ văn bản trong textbox
        self.textbox.delete(
            index1="0.0",
            index2="end"
        )
         # Thêm văn bản mới vào đầu textbox
        self.textbox.insert("0.0", text= text)

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def button_xem_nhan_vien(self):
        # region Lấy danh sách công việc của nhân viên đầu tiên để hiển thị
        self.nhan_vien : NhanVien = self.data.get_thong_tin_nhan_vien(
            nam = self.nam_optionemenu.get(),
            thang= self.thang_optionemenu.get(),
            ma_nhan_vien= self.nhan_vien_optionemenu.get()
        )

        # region Update checkbox cho nhân viên
        lst_nhan_vien = self.data.get_danh_sach_nhan_vien(
                nam= self.nam_optionemenu.get(),
                thang= self.thang_optionemenu.get()
            )
        self.dict_nhan_vien_checkbox = {}
        for idx, x in enumerate(lst_nhan_vien, start= 1):
            self.dict_nhan_vien_checkbox[x] = customtkinter.CTkCheckBox(master=self.sidebar_frame_print, text = x  )
            if idx % 2 == 0: #chan
                self.dict_nhan_vien_checkbox[x].grid(row=idx, column=0, pady=(0, 0), padx=0, sticky="n")
            else:
                self.dict_nhan_vien_checkbox[x].grid(row=idx+1, column=1, pady=(0, 0), padx=0, sticky="n")

        # endregion
        self.ho_ten_nv = self.nhan_vien.get_ho_ten()
        self.ma_so_nv = self.nhan_vien.get_msnv()
        self.cac_cong_viec_trong_thang = self.nhan_vien.get_cac_cong_viec_trong_thang()
        self.so_cong_viec_trong_thang = self.nhan_vien.get_so_cong_viec_trong_thang()
        self.so_gio_cong = self.nhan_vien.get_so_gio_cong()
        self.so_phut_tre = self.nhan_vien.get_so_phut_di_tre()
        self.so_gio_ve_som, *_ = self.nhan_vien.get_so_gio_ve_som()
        self.so_gio_tang_ca, *_= self.nhan_vien.get_so_gio_tang_ca()
        self.so_gio_nghi_phep, *_ = self.nhan_vien.get_so_gio_nghi_phep()

        lst_cong_viec = self.cac_cong_viec_trong_thang + [TANG_CA]
        self.cac_cong_viec_optionemenu.configure(
            values = lst_cong_viec
        )
        # endregion
        self.ho_ten_label.configure(text=f"Họ tên: {self.ho_ten_nv}")
        self.so_gio_cong_label.configure(text=f"Số ngày công: {self.so_gio_cong} (giờ)")
        self.tang_ca_label.configure(text=f"Tăng ca: {self.so_gio_tang_ca} (giờ)")
        self.tre_label.configure(text=f"Trễ: {self.so_phut_tre} (phút)")
        self.som_label.configure(text=f"Sớm: {self.so_gio_ve_som} (giờ)")
        self.phep_label.configure(text=f"Phép: {self.so_gio_nghi_phep} (giờ)")

    # def update_partial(self):
    #     # Lấy tất cả các file trong nằm trong cây thư mục
    #     lst_path = []
    #     for fd in os.listdir(DATAPATH):
    #         fd_path = os.path.join(DATAPATH, fd)
    #         for fn in os.listdir(fd_path):
    #             path = os.path.join(fd_path, fn)
    #             lst_path.append(path)
    #     # Tìm ra file sửa trong cây thư mục
    #     top_3_change = sorted(lst_path, key = latest_change)[:3]
    #     # cập nhật dữ liệu
    #     self.data.update(top_3_change)
    #     dump_pickle(fn = DATANAME, obj= self.data)
    #     lst_nam = self.data.get_danh_sach_nam()
    #     self.nam_optionemenu.configure(
    #         values = sorted(lst_nam)
    #     )
    #     lst_thang = self.data.get_danh_sach_thang(nam = self.nam_optionemenu.get())
    #     self.thang_optionemenu.configure(
    #         values = sorted(lst_thang, key = lambda x: int(x))
    #     )

    def update_whole(self):
        self.data=ToanBoData(update_whole= True)
        dump_pickle(fn = DATANAME, obj= self.data)
        lst_nam = self.data.get_danh_sach_nam()
        self.nam_optionemenu.configure(
            values = sorted(lst_nam)
        )

        lst_thang = self.data.get_danh_sach_thang(nam = self.nam_optionemenu.get())
        self.thang_optionemenu.configure(
            values = sorted(lst_thang, key = lambda x: int(x))
        )

    def upload_file(self):
        file_path = askopenfile(mode='r', filetypes=[('Excel_97', '*xls'), ('Excel_97', '*xlsx')])
        if file_path is not None:
            pass
        print(f'file_path: {file_path.name}')
        fn = file_path.name.replace('/', '\\')
        self.data.update([fn])
        print(f'Cập nhật thành công file {fn}')
        dump_pickle(fn = DATANAME, obj= self.data)
        lst_nam = self.data.get_danh_sach_nam()
        self.nam_optionemenu.configure(
            values = sorted(lst_nam)
        )

        lst_thang = self.data.get_danh_sach_thang(nam = self.nam_optionemenu.get())
        self.thang_optionemenu.configure(
            values = sorted(lst_thang, key = lambda x: int(x))
        )
    
    def change_month_option(self, year):
        lst_thang = self.data.get_danh_sach_thang(nam = year)
        self.thang_optionemenu.configure(values = sorted(lst_thang, key = lambda x: int(x)))

if __name__ == "__main__":
    app = App()
    app.mainloop()
    # data = load_pickle(fn = 'full_data_v3.plk' )
    # print(data)